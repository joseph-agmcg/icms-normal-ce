"""
Extração de dados de planilhas Excel de filiais (formato ICMS).
Captura todos os dados: todas as colunas do cabeçalho e todas as linhas até a linha de TOTAL.
O período de referência é lido da área do título (ao lado de "APURAÇÃO DE ICMS CEARA"), não de uma coluna.
"""

import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from sefaz_ce.logger import configurar_logger_da_aplicacao

logger = configurar_logger_da_aplicacao(__name__)

# Nome exato da coluna I.E. no Excel (estilo icms.xlsx). Outras variações aceitas abaixo.
NOME_COLUNA_IE = "INSC.ESTADUAL"

# Outros nomes aceitos para a coluna I.E. (planilhas em outros formatos)
SINONIMOS_IE = (
    "INSC.ESTADUAL",
    "INSC. ESTADUAL",
    "i.e.",
    "ie",
    "i..e.",
    "inscrição estadual",
    "inscricao estadual",
)

# Número máximo de linhas a varrer para encontrar o cabeçalho
MAX_LINHAS_BUSCA_CABECALHO = 25

# Linhas e colunas da área do título onde pode estar o período (ex.: 1/1/2026 ou jan-26)
MAX_LINHAS_BUSCA_PERIODO = 12
MAX_COLUNAS_BUSCA_PERIODO = 30

# Meses em português (formato jan-26, jan/26, etc.)
MESES_ABREV = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

# Número de dígitos da I.E. no Ceará (preenchimento com zeros à esquerda quando faltar)
DIGITOS_IE_CE = 9


def _normalizar_cabecalho(texto: str | None) -> str:
    """Retorna o texto em minúsculo, sem acentos, sem pontos e com espaços normais."""
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = "".join(c for c in s if c != ".")
    return " ".join(s.split())


def _celula_bate_nome_ie(valor: str | None) -> bool:
    """Verifica se o valor da célula corresponde ao nome da coluna I.E. (ou sinônimos)."""
    n = _normalizar_cabecalho(valor)
    if not n:
        return False
    for nome in SINONIMOS_IE:
        if _normalizar_cabecalho(nome) == n:
            return True
    if n == "inscestadual" or n == "insc estadual":
        return True
    return False


def _encontrar_linha_cabecalho(planilha: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    """Retorna o índice 0-based da linha que contém a coluna I.E. (cabeçalho)."""
    for indice_linha in range(MAX_LINHAS_BUSCA_CABECALHO):
        row_num = indice_linha + 1
        linha = list(planilha.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        if not linha:
            continue
        celulas = linha[0]
        for valor in celulas:
            if _celula_bate_nome_ie(valor):
                return indice_linha
    return None


def _mapear_cabecalho(planilha: openpyxl.worksheet.worksheet.Worksheet, indice_linha: int) -> tuple[dict[str, int], int]:
    """
    Lê a linha de cabeçalho e retorna (nome_coluna -> índice_coluna, índice_coluna_ie).
    Todas as colunas são mapeadas; colunas sem nome recebem "col_N". Chaves únicas.
    """
    row_num = indice_linha + 1
    linha = list(planilha.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    celulas = list(linha[0]) if linha else []
    nome_para_indice: dict[str, int] = {}
    coluna_ie: int | None = None
    for indice_col, valor in enumerate(celulas):
        nome = valor
        if nome is None or (isinstance(nome, str) and not nome.strip()):
            nome = f"col_{indice_col}"
        else:
            nome = str(nome).strip()
        if nome in nome_para_indice:
            nome = f"{nome}_{indice_col}"
        nome_para_indice[nome] = indice_col
        if _celula_bate_nome_ie(valor):
            coluna_ie = indice_col
    if coluna_ie is None:
        coluna_ie = -1
    return nome_para_indice, coluna_ie


def _extrair_periodo_da_area_titulo(planilha: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int]:
    """
    Procura o período de referência na área do título da planilha (ao lado de "APURAÇÃO DE ICMS CEARA").
    Varre as primeiras linhas e colunas. Aceita:
    - Data: 1/1/2026, 01/01/2026 (dia/mês/ano ou mês/ano)
    - Mês/ano formatado: jan-26, jan/26, jan 26
    - Célula com valor datetime ou número serial do Excel.
    Levanta ValueError se não encontrar (sem fallback).
    """
    for row_num in range(1, MAX_LINHAS_BUSCA_PERIODO + 1):
        try:
            row = list(planilha.iter_rows(min_row=row_num, max_row=row_num, max_col=MAX_COLUNAS_BUSCA_PERIODO, values_only=True))
        except Exception:
            continue
        if not row:
            continue
        celulas = list(row[0]) if row else []
        for valor in celulas:
            if valor is None:
                continue
            mes, ano = _parsear_celula_periodo(valor)
            if mes is not None and ano is not None:
                logger.info("Período de referência encontrado na área do título: %02d/%d (linha %d).", mes, ano, row_num)
                return mes, ano
    raise ValueError(
        "Período de referência não encontrado na área do título. "
        "Inclua a data (ex.: 1/1/2026) ou mês/ano (ex.: jan-26) ao lado do título 'APURAÇÃO DE ICMS CEARA'."
    )


def _parsear_celula_periodo(valor: object) -> tuple[int | None, int | None]:
    """Tenta extrair (mês, ano) de uma célula: datetime, número serial Excel, ou string (1/1/2026, jan-26)."""
    if isinstance(valor, datetime):
        if 1 <= valor.month <= 12 and 1995 <= valor.year <= 2030:
            return valor.month, valor.year
        return None, None
    if isinstance(valor, (int, float)):
        # Número serial de data do Excel (1 = 1900-01-01)
        try:
            if 2000 < valor < 100000:  # faixa razoável para datas
                d = datetime(1899, 12, 30) + timedelta(days=int(valor))
                if 1 <= d.month <= 12 and 1995 <= d.year <= 2030:
                    return d.month, d.year
        except (ValueError, OverflowError):
            pass
        return None, None
    if isinstance(valor, str):
        s = valor.strip()
        if not s:
            return None, None
        # Formato "jan-26" ou "jan/26" ou "jan 26"
        m = re.match(r"^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[\s\-/](\d{2})$", s.lower())
        if m:
            mes_str, ano_2 = m.group(1), m.group(2)
            mes = MESES_ABREV.get(mes_str)
            ano = 2000 + int(ano_2) if int(ano_2) < 50 else 1900 + int(ano_2)
            if mes and 1995 <= ano <= 2030:
                return mes, ano
        # Formato "1/1/2026" ou "1/2026"
        part = [p.strip() for p in re.split(r"[/\-.]", s) if p.strip()]
        if len(part) >= 2:
            try:
                nums = [int(p) for p in part if p.isdigit()]
                if len(nums) >= 2:
                    if len(nums) == 2:
                        mes, ano = nums[0], nums[1]
                    else:
                        mes, ano = nums[1], nums[2]
                    if 1 <= mes <= 12 and 1995 <= ano <= 2030:
                        return mes, ano
            except (ValueError, IndexError):
                pass
    return None, None


def _valor_celula_para_python(valor: object) -> object:
    """Mantém número como número, texto como string; None permanece None."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return valor
    return str(valor).strip() if str(valor).strip() else None


def _valor_ie_para_string(valor: str | int | float | None) -> str | None:
    """Converte o valor da célula I.E. para string (pode vir número ou texto)."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        s = str(int(valor)).strip()
        return _normalizar_ie_zeros_esquerda(s) if s else None
    s = str(valor).strip()
    return _normalizar_ie_zeros_esquerda(s) if s else None


def _normalizar_ie_zeros_esquerda(ie: str) -> str:
    """
    Se a I.E. for só dígitos e tiver menos que DIGITOS_IE_CE caracteres,
    preenche com zeros à esquerda (evita perder o zero inicial vindo do Excel como número).
    """
    if not ie or len(ie) >= DIGITOS_IE_CE:
        return ie
    if not ie.isdigit():
        return ie
    return ie.zfill(DIGITOS_IE_CE)


def _eh_linha_fim_dados(valor_ie: object) -> bool:
    """Indica se o valor na coluna I.E. é total/footer (parar de ler)."""
    if valor_ie is None:
        return False
    v = str(valor_ie).strip().upper()
    if not v:
        return False
    return v.startswith("TOTAL") or v == "CEARA"


def extrair_todos_os_dados(caminho_arquivo: str | Path) -> tuple[list[dict[str, object]], dict[str, int], int, int]:
    """
    Lê o Excel e retorna (lista de linhas como dict, mapeamento nome_coluna -> índice, mes_ref, ano_ref).
    O período de referência é lido da área do título (ao lado de "APURAÇÃO DE ICMS CEARA"), não de uma coluna.
    """
    caminho = Path(caminho_arquivo)
    logger.info("Iniciando extração do Excel: %s", caminho.resolve())

    if not caminho.exists():
        logger.error("Arquivo não encontrado: %s", caminho)
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    logger.debug("Abrindo workbook (read_only=True, data_only=True).")
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Nenhuma planilha ativa no arquivo.")
        logger.debug("Planilha ativa: %s", ws.title)

        logger.debug("Buscando linha do cabeçalho (máx. %d linhas).", MAX_LINHAS_BUSCA_CABECALHO)
        indice_cabecalho = _encontrar_linha_cabecalho(ws)
        if indice_cabecalho is None:
            raise ValueError(
                "Coluna de Inscrição Estadual (I.E.) não encontrada no cabeçalho. "
                "Use um dos nomes: INSC.ESTADUAL, INSC. ESTADUAL, I.E., etc."
            )
        logger.info("Cabeçalho encontrado na linha %d (1-based: %d).", indice_cabecalho, indice_cabecalho + 1)

        nome_para_indice, coluna_ie = _mapear_cabecalho(ws, indice_cabecalho)
        colunas_ordenadas = [k for k, _ in sorted(nome_para_indice.items(), key=lambda x: x[1])]
        logger.info("Colunas mapeadas (%d): %s", len(colunas_ordenadas), colunas_ordenadas)
        logger.debug("Índice da coluna I.E.: %d", coluna_ie)

        if coluna_ie < 0:
            raise ValueError("Coluna I.E. não identificada no mapeamento do cabeçalho.")

        mes_ref, ano_ref = _extrair_periodo_da_area_titulo(ws)

        min_row = indice_cabecalho + 2
        linhas: list[dict[str, object]] = []
        num_linha_lida = 0

        for row in ws.iter_rows(min_row=min_row, values_only=True):
            celulas = list(row)
            valor_ie = celulas[coluna_ie] if coluna_ie < len(celulas) else None
            if _eh_linha_fim_dados(valor_ie):
                logger.debug("Linha de fim de dados detectada (I.E.=%s). Parando leitura.", valor_ie)
                break
            linha_dict: dict[str, object] = {}
            for nome_col, idx in nome_para_indice.items():
                if idx < len(celulas):
                    linha_dict[nome_col] = _valor_celula_para_python(celulas[idx])
                else:
                    linha_dict[nome_col] = None
            # Normaliza a coluna I.E. com zeros à esquerda (Excel pode retornar número sem o zero)
            ie_key = _obter_chave_ie(nome_para_indice)
            if ie_key is not None and linha_dict.get(ie_key) is not None:
                normalizada = _valor_ie_para_string(linha_dict[ie_key])
                if normalizada is not None:
                    linha_dict[ie_key] = normalizada
            linhas.append(linha_dict)
            num_linha_lida += 1

        logger.info("Linhas de dados lidas: %d (antes da linha TOTAL).", len(linhas))
        if linhas:
            logger.debug("Primeira linha de dados: %s", linhas[0])
            if len(linhas) > 1:
                logger.debug("Última linha de dados: %s", linhas[-1])

        return linhas, nome_para_indice, mes_ref, ano_ref
    finally:
        wb.close()
        logger.debug("Workbook fechado.")


def _obter_chave_ie(nome_para_indice: dict[str, int]) -> str | None:
    """Retorna a chave do cabeçalho que corresponde à coluna I.E."""
    for chave in nome_para_indice:
        if _celula_bate_nome_ie(chave):
            return chave
    return None


def _obter_chave_normal(nome_para_indice: dict[str, int]) -> str | None:
    """Retorna a chave do cabeçalho que corresponde à coluna NORMAL (valor principal ICMS)."""
    n = _normalizar_cabecalho("NORMAL")
    for chave in nome_para_indice:
        if _normalizar_cabecalho(chave) == n:
            return chave
    return None


def _obter_chave_total(nome_para_indice: dict[str, int]) -> str | None:
    """Retorna a chave do cabeçalho que corresponde à coluna TOTAL (valor principal para DAE)."""
    n = _normalizar_cabecalho("TOTAL")
    for chave in nome_para_indice:
        if _normalizar_cabecalho(chave) == n:
            return chave
    return None


def ie_apenas_digitos(ie: str) -> str:
    """
    Retorna a I.E. apenas com dígitos, com 9 caracteres (zeros à esquerda).
    Usado para preencher o campo no site (sem pontos nem traços).
    """
    if not ie:
        return ""
    digitos = "".join(c for c in str(ie).strip() if c.isdigit())
    return digitos.zfill(DIGITOS_IE_CE) if len(digitos) <= DIGITOS_IE_CE else digitos[:DIGITOS_IE_CE]


def obter_dados_para_dae(
    linhas: list[dict[str, object]],
    nome_para_indice: dict[str, int],
    mes_ref: int | None = None,
    ano_ref: int | None = None,
) -> list[dict[str, object]]:
    """
    Retorna uma lista de dict por linha, com: ie, ie_digitos, valor_normal, mes_ref, ano_ref.
    O valor principal (valor_normal) é lido da coluna TOTAL da planilha.
    mes_ref e ano_ref são obrigatórios (vêm da extração); se None, levanta ValueError.
    """
    if mes_ref is None or ano_ref is None:
        raise ValueError("Período de referência (mes_ref/ano_ref) é obrigatório. Verifique a extração da planilha.")
    ie_key = _obter_chave_ie(nome_para_indice)
    total_key = _obter_chave_total(nome_para_indice)
    if ie_key is None:
        raise ValueError("Coluna I.E. não encontrada no mapeamento. Nenhum dado para DAE.")
    resultado: list[dict[str, object]] = []
    for row in linhas:
        valor_ie = row.get(ie_key)
        ie_str = _valor_ie_para_string(valor_ie) if isinstance(valor_ie, (str, int, float)) else (str(valor_ie) if valor_ie is not None else None)
        if not ie_str:
            continue
        valor_normal: float | None = None
        if total_key:
            v = row.get(total_key)
            if isinstance(v, (int, float)):
                valor_normal = float(v)
            elif v is not None and str(v).strip():
                try:
                    valor_normal = float(str(v).strip().replace(",", "."))
                except ValueError:
                    valor_normal = None
        resultado.append({
            "ie": ie_str,
            "ie_digitos": ie_apenas_digitos(ie_str),
            "valor_normal": valor_normal,
            "mes_ref": mes_ref,
            "ano_ref": ano_ref,
        })
    logger.info("Dados para DAE: %d registro(s) (período %02d/%d).", len(resultado), mes_ref, ano_ref)
    return resultado


def obter_ies_dos_dados(
    linhas: list[dict[str, object]],
    nome_para_indice: dict[str, int],
) -> list[str]:
    """
    Extrai a lista de I.E. a partir dos dados já carregados (evita reler o arquivo).
    """
    ie_key = _obter_chave_ie(nome_para_indice)
    if ie_key is None:
        return []
    ies: list[str] = []
    for row in linhas:
        valor = row.get(ie_key)
        ie = _valor_ie_para_string(valor) if isinstance(valor, (str, int, float)) else (str(valor) if valor is not None else None)
        if ie:
            ies.append(ie)
    return ies


def extrair_ies_do_excel(caminho_arquivo: str | Path) -> list[str]:
    """
    Lê o arquivo Excel e retorna uma lista com todos os valores da coluna I.E.
    Usa a mesma extração completa (todos os dados); retorna apenas o campo I.E. de cada linha.
    """
    linhas, nome_para_indice, _, _ = extrair_todos_os_dados(caminho_arquivo)
    ie_key = _obter_chave_ie(nome_para_indice)
    if ie_key is None:
        logger.warning("Chave I.E. não encontrada no mapeamento. Nenhuma I.E. extraída.")
        return []
    ies: list[str] = []
    for row in linhas:
        valor = row.get(ie_key)
        ie = _valor_ie_para_string(valor) if isinstance(valor, (str, int, float)) else (str(valor) if valor is not None else None)
        if ie:
            ies.append(ie)
    logger.info("I.E.s extraídas: %d. Valores: %s", len(ies), ies if len(ies) <= 20 else ies[:20] + [f"... e mais {len(ies) - 20}"])
    return ies

